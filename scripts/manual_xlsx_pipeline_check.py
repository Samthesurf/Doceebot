from __future__ import annotations

import json
import shutil
from pathlib import Path
from tempfile import TemporaryDirectory

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from whatsapp_ai_agent.config import Settings, get_settings
from whatsapp_ai_agent.db.models import Base, Membership, Organization, User
from whatsapp_ai_agent.db.session import get_db_session
from whatsapp_ai_agent.main import create_app

HEADERS = [
    "Date",
    "Start Time",
    "End Time",
    "Activity",
    "People Participated",
    "Site",
    "Status",
    "Notes",
]

INITIAL_ROWS = [
    {
        "Date": "2026-06-10",
        "Start Time": "09:00",
        "End Time": "11:30",
        "Activity": "Initial site inspection",
        "People Participated": "Samuel, Ada",
        "Site": "Main Plant",
        "Status": "done",
        "Notes": "Baseline inspection completed",
    },
    {
        "Date": "2026-06-14",
        "Start Time": "10:00",
        "End Time": "15:00",
        "Activity": "Cable tray installation",
        "People Participated": "Samuel, Tunde",
        "Site": "Warehouse",
        "Status": "in_progress",
        "Notes": "Covers pending",
    },
]

SCENARIOS = [
    {
        "name": "append_non_consecutive_date_1",
        "sid": "SMPIPE001",
        "body": (
            "Update the Daily Activity Log Excel. On 2026-07-02 from 08:15 to 10:45, "
            "activity was generator room inspection at Main Plant. People participated: "
            "Samuel and Musa. Status done. Notes: no fault found."
        ),
        "expected": {
            "Date": "2026-07-02",
            "Start Time": "08:15",
            "End Time": "10:45",
            "Activity": "generator room inspection",
            "People Participated": "Samuel and Musa",
            "Site": "Main Plant",
            "Status": "done",
            "Notes": "no fault found",
        },
    },
    {
        "name": "append_non_consecutive_date_2",
        "sid": "SMPIPE002",
        "body": (
            "Append to the Daily Activity Log: On 2026-07-05, 14:00 to 16:30, "
            "activity was inverter battery terminal cleaning at Inverter Room. People: "
            "Ada and Samuel. Status done. Notes: tightened loose lugs."
        ),
        "expected": {
            "Date": "2026-07-05",
            "Start Time": "14:00",
            "End Time": "16:30",
            "Activity": "inverter battery terminal cleaning",
            "People Participated": "Ada and Samuel",
            "Site": "Inverter Room",
            "Status": "done",
            "Notes": "tightened loose lugs",
        },
    },
    {
        "name": "append_multi_people",
        "sid": "SMPIPE003",
        "body": (
            "Add this to the Daily Activity Log Excel: 2026-07-12 from 09:30 to 13:10, "
            "activity: DB dressing and circuit labeling, people participated: Samuel, John, "
            "Grace, site: Admin Block, status done."
        ),
        "expected": {
            "Date": "2026-07-12",
            "Start Time": "09:30",
            "End Time": "13:10",
            "Activity": "DB dressing and circuit labeling",
            "People Participated": "Samuel, John, Grace",
            "Site": "Admin Block",
            "Status": "done",
        },
    },
    {
        "name": "append_later_month",
        "sid": "SMPIPE004",
        "body": (
            "Record in Daily Activity Log: on 2026-08-03 from 11:20 to 12:05, "
            "activity: tested emergency lighting circuit, people participated: Samuel only, "
            "site: Workshop, status done, notes: two fittings weak."
        ),
        "expected": {
            "Date": "2026-08-03",
            "Start Time": "11:20",
            "End Time": "12:05",
            "Activity": "tested emergency lighting circuit",
            "People Participated": "Samuel only",
            "Site": "Workshop",
            "Status": "done",
            "Notes": "two fittings weak",
        },
    },
    {
        "name": "update_existing_preloaded_row",
        "sid": "SMPIPE005",
        "body": (
            "Update existing Daily Activity Log row for 2026-06-14, 10:00 to 15:00: "
            "activity remains cable tray installation, people participated: Samuel, "
            "Tunde and Kemi, site: Warehouse, status done, notes: added missing "
            "trunking covers."
        ),
        "expected": {
            "Date": "2026-06-14",
            "Start Time": "10:00",
            "End Time": "15:00",
            "Activity": "cable tray installation",
            "People Participated": "Samuel, Tunde and Kemi",
            "Site": "Warehouse",
            "Status": "done",
            "Notes": "added missing trunking covers",
        },
        "must_update_existing": True,
    },
]


def create_initial_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily Log"
    sheet.append(HEADERS)
    for row in INITIAL_ROWS:
        sheet.append([row.get(header) for header in HEADERS])
    workbook.save(path)


def read_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [str(cell.value) for cell in sheet[1]]
    rows: list[dict[str, object]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if any(value is not None for value in values):
            rows.append({header: value for header, value in zip(headers, values, strict=False)})
    return headers, rows


def norm(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def row_matches(row: dict[str, object], expected: dict[str, object]) -> bool:
    for key, expected_value in expected.items():
        expected_norm = norm(expected_value)
        actual_norm = norm(row.get(key))
        if expected_norm and not actual_norm:
            return False
        if expected_norm not in actual_norm and actual_norm not in expected_norm:
            return False
    return True


def find_best_date_row(rows: list[dict[str, object]], date: str, start_time: str | None = None):
    candidates = [row for row in rows if norm(row.get("Date")) == norm(date)]
    if start_time:
        exact = [row for row in candidates if norm(row.get("Start Time")) == norm(start_time)]
        if exact:
            return exact[0]
    return candidates[0] if candidates else None


def main() -> None:
    with TemporaryDirectory(prefix="doceebot-xlsx-pipeline-") as tmp:
        tmp_path = Path(tmp)
        storage_dir = tmp_path / "storage"
        input_path = tmp_path / "daily_activity_log_seed.xlsx"
        create_initial_workbook(input_path)

        db_path = tmp_path / "pipeline.sqlite"
        engine = create_engine(f"sqlite+pysqlite:///{db_path}")
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine, autoflush=False, autocommit=False)
        db_session = Session()

        org = Organization(name="Pipeline Test Org")
        user = User(display_name="Samuel Test Worker", phone_number="+2348012345678")
        db_session.add_all([org, user])
        db_session.flush()
        db_session.add(Membership(org_id=org.id, user_id=user.id, role="worker"))
        db_session.commit()

        settings = Settings(
            app_env="development",
            local_storage_dir=str(storage_dir),
            media_storage_backend="local",
            twilio_webhook_auth_enabled=False,
            _env_file=".env",
        )
        app = create_app(settings)

        def override_db_session():
            try:
                yield db_session
            finally:
                pass

        app.dependency_overrides[get_db_session] = override_db_session
        app.dependency_overrides[get_settings] = lambda: settings

        client = TestClient(app)
        with input_path.open("rb") as handle:
            upload_response = client.post(
                "/dashboard/documents/upload",
                data={
                    "org_id": str(org.id),
                    "owner_user_id": str(user.id),
                    "display_name": "Daily Activity Log",
                    "summary": (
                        "Daily activity log workbook with columns: Date, Start Time, End Time, "
                        "Activity, People Participated, Site, Status, Notes."
                    ),
                    "tags": "daily-activity-log,excel,work-log",
                },
                files={
                    "file": (
                        "daily_activity_log.xlsx",
                        handle,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        upload_response.raise_for_status()
        document = upload_response.json()

        scenario_results = []
        for scenario in SCENARIOS:
            response = client.post(
                "/webhooks/twilio/whatsapp",
                data={
                    "MessageSid": scenario["sid"],
                    "From": "whatsapp:+2348012345678",
                    "WaId": "2348012345678",
                    "Body": scenario["body"],
                    "NumMedia": "0",
                },
            )
            scenario_results.append(
                {
                    "name": scenario["name"],
                    "status_code": response.status_code,
                    "twiml": response.text,
                }
            )
            response.raise_for_status()

        # Upload returns the same storage key that table updates overwrite.
        final_path = storage_dir / document["storage_key"]
        headers, rows = read_rows(final_path)
        failures = []

        if headers != HEADERS:
            failures.append({"type": "headers", "expected": HEADERS, "actual": headers})

        expected_row_count = len(INITIAL_ROWS) + 4
        if len(rows) != expected_row_count:
            failures.append(
                {"type": "row_count", "expected": expected_row_count, "actual": len(rows)}
            )

        unchanged = find_best_date_row(rows, "2026-06-10", "09:00")
        if unchanged is None or not row_matches(unchanged, INITIAL_ROWS[0]):
            failures.append(
                {
                    "type": "preexisting_unchanged",
                    "expected": INITIAL_ROWS[0],
                    "actual": unchanged,
                }
            )

        date_counts: dict[str, int] = {}
        for row in rows:
            date_counts[str(row.get("Date"))] = date_counts.get(str(row.get("Date")), 0) + 1
        if date_counts.get("2026-06-14") != 1:
            failures.append({"type": "existing_row_duplicated", "date_counts": date_counts})

        actual_checks = []
        for scenario in SCENARIOS:
            expected = scenario["expected"]
            actual = find_best_date_row(rows, expected["Date"], expected.get("Start Time"))
            ok = bool(actual and row_matches(actual, expected))
            actual_checks.append(
                {
                    "name": scenario["name"],
                    "ok": ok,
                    "expected": expected,
                    "actual": actual,
                }
            )
            if not ok:
                failures.append(actual_checks[-1])

        artifact_dir = Path("/root/Doceebot/tmp_pipeline_artifacts")
        artifact_dir.mkdir(parents=True, exist_ok=True)
        final_artifact = artifact_dir / "daily_activity_log_after_api_pipeline.xlsx"
        seed_artifact = artifact_dir / "daily_activity_log_before_api_pipeline.xlsx"
        shutil.copyfile(final_path, final_artifact)
        shutil.copyfile(input_path, seed_artifact)

        report = {
            "document_id": document["id"],
            "seed_artifact": str(seed_artifact),
            "final_artifact": str(final_artifact),
            "scenario_responses": scenario_results,
            "headers": headers,
            "rows": rows,
            "checks": actual_checks,
            "failures": failures,
            "passed": not failures,
        }
        print(json.dumps(report, indent=2, default=str))
        if failures:
            raise SystemExit(1)


if __name__ == "__main__":
    main()
